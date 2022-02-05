import myLib
import openpyxl
import datetime

# strTime = '2019-07-11 11:03'  # 给定一个时间，此是个字符串
# startTime = datetime.datetime.strptime(strTime, "%Y-%m-%d %H:%M")  # 把strTime转化为时间格式,后面的秒位自动补位的
#
# wb = openpyxl.load_workbook('BasalPlate/21财9班学生体温检测表.xlsx')
#
# # print(ws.cell(column=1,row=1).value)
# time = datetime.datetime.strptime('2022-01-14','%Y-%m-%d')
# # timeP = datetime.datetime.strptime('2022-01-15','%Y-%m-%d')
# # print(type(timeP - time))
# value = 36.6
# temp = myLib.writeSheetDateFrom0114(wb,time,value)
# print(temp)


wb = openpyxl.load_workbook('BasalPlate/21财9班学生体温检测表.xlsx')
basalPlate = '21财9班学生体温检测表'
startTime = datetime.datetime.strptime('2022-01-14','%Y-%m-%d')
endTime = datetime.datetime.strptime('2022-01-20','%Y-%m-%d')
temp = myLib.saveFileByTimeToOutput(basalPlate,startTime,endTime)
print(temp)