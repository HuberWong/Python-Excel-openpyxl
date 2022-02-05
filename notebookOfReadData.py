"""
    读取
    openpyxl
"""

import openpyxl as opx

wb = opx.load_workbook('./BasalPlate/20财14班学生体温检测表.xlsx')

print(wb.sheetnames)

for sheet in wb:
    print(sheet.title)

mySheet = wb.create_sheet('mySheet')
print(wb.sheetnames)

# sheet3 = wb.get_sheet_by_name('Sheet3')

ws = wb.active
print(ws)
print(ws['A1'])
print(ws['A1'].value)

c = ws['B1']
print('Row {}, Column {} is {}'.format(c.row,c.column,c.value))
print('Row',c.row,'Column',c.column,'is',c.value)

wb.save('./TestFolder/output.xlsx')

print(ws.cell(row=1,column=2))
print(ws.cell(row=1,column=2).value)

for i in range (1,8,2):
    print(i,ws.cell(row=i,column=2).value)

colC = ws['B']
print(colC)
print(colC[2].value) # 索引从零开始
col_range = ws['A:B']
row_range = ws[2:6]

for col in col_range:
    for cell in col:
        print(cell.value)

for row in row_range:
    for cell in row:
        print(cell.value)

for row in ws.iter_rows(min_row=1,max_row=2,max_col=2):
    for cell in row:
        print(cell)

print(tuple(ws.rows))

cell_range = ws['A1:B10']
for rowOfCellObject in cell_range:
    for cellObj in rowOfCellObject:
        print(cellObj.coordinate, cell.value)
    print('------ End of Row ------')

print(ws.max_row,'*',ws.max_column)

from openpyxl.utils import get_column_letter,column_index_from_string
print(get_column_letter(2),get_column_letter(47),get_column_letter(900))
print(column_index_from_string('AHH'))
